#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保险费用Excel处理工具 v4.0 (优化版)
功能：
1. 季度处理：合并数据、计算汇率、生成季度保费、分行统计
2. 年度对比：对比四季度累加保费与年初总表，计算差额
3. 性能优化：向量化操作、批量写入、多线程处理
4. 新增功能：配置保存、批量处理、自动汇率、历史记录
"""

import sys
import os
import json
from concurrent.futures import ThreadPoolExecutor
import numpy as np

# 修复 macOS Qt 平台插件问题
def _fix_qt_plugin():
    try:
        import PyQt5
        for subdir in ['Qt5/plugins', 'Qt/plugins']:
            path = os.path.join(os.path.dirname(PyQt5.__file__), subdir)
            if os.path.exists(path):
                os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = path
                break
    except: pass
_fix_qt_plugin()

from datetime import datetime, date
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QProgressBar, QTextEdit,
    QGroupBox, QDoubleSpinBox, QMessageBox, QFrame, QComboBox,
    QGridLayout, QLineEdit, QTabWidget, QCheckBox, QSpinBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl, QSettings
from PyQt5.QtGui import QFont, QDesktopServices

# 配置文件路径
CONFIG_FILE = os.path.expanduser("~/.insurance_processor_config.json")

# 样式表
STYLE_SHEET = """
QMainWindow {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #1a1a2e, stop:0.5 #16213e, stop:1 #0f3460);
}
QWidget {
    font-family: 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif;
    font-size: 13px;
    color: #e8e8e8;
}
QTabWidget::pane {
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-radius: 8px;
    background: rgba(0, 0, 0, 0.2);
    margin-top: -1px;
}
QTabBar::tab {
    background: rgba(138, 43, 226, 0.2);
    color: #9ca3af;
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-bottom: none;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    padding: 10px 25px;
    margin-right: 4px;
    font-weight: bold;
}
QTabBar::tab:selected {
    background: rgba(138, 43, 226, 0.5);
    color: #ffffff;
}
QTabBar::tab:hover:!selected {
    background: rgba(138, 43, 226, 0.35);
    color: #e8e8e8;
}
QGroupBox {
    background: rgba(255, 255, 255, 0.05);
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-radius: 10px;
    margin-top: 15px;
    padding: 15px;
    font-weight: bold;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 15px;
    padding: 0 8px;
    color: #bb86fc;
}
QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #7c3aed, stop:1 #a855f7);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 20px;
    font-weight: bold;
    min-width: 100px;
}
QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #8b5cf6, stop:1 #c084fc);
}
QPushButton:pressed {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #6d28d9, stop:1 #9333ea);
}
QPushButton:disabled {
    background: #4a4a6a;
    color: #888;
}
QPushButton#successBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #059669, stop:1 #10b981);
}
QPushButton#successBtn:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #10b981, stop:1 #34d399);
}
QPushButton#compareBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #2563eb, stop:1 #3b82f6);
}
QPushButton#compareBtn:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #3b82f6, stop:1 #60a5fa);
}
QLineEdit, QDoubleSpinBox, QComboBox {
    background: rgba(0, 0, 0, 0.3);
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-radius: 6px;
    padding: 8px 12px;
    color: #e8e8e8;
}
QLineEdit:hover, QDoubleSpinBox:hover, QComboBox:hover {
    border-color: #bb86fc;
}
QLineEdit:focus, QDoubleSpinBox:focus {
    border-color: #a855f7;
    background: rgba(0, 0, 0, 0.4);
}
QComboBox::drop-down { border: none; width: 30px; }
QComboBox QAbstractItemView {
    background: #1a1a2e;
    border: 1px solid rgba(138, 43, 226, 0.5);
    selection-background-color: rgba(138, 43, 226, 0.4);
}
QDoubleSpinBox::up-arrow {
    width: 0;
    height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-bottom: 5px solid #e8e8e8; /* Visible arrow color */
}
QDoubleSpinBox::down-arrow {
    width: 0;
    height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #e8e8e8; /* Visible arrow color */
}
QProgressBar {
    background: rgba(0, 0, 0, 0.3);
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-radius: 8px;
    text-align: center;
    color: white;
    height: 25px;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #7c3aed, stop:1 #a855f7);
    border-radius: 7px;
}
QTextEdit {
    background: rgba(0, 0, 0, 0.4);
    border: 1px solid rgba(138, 43, 226, 0.3);
    border-radius: 8px;
    padding: 10px;
    font-family: 'Monaco', 'Consolas', monospace;
    font-size: 12px;
}
QLabel#titleLabel { font-size: 22px; font-weight: bold; color: #bb86fc; }
QLabel#subtitleLabel { font-size: 13px; color: #9ca3af; }
"""


class ConfigManager:
    """配置管理器"""
    @staticmethod
    def load_config():
        """加载配置"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {
            'prev_usd_rate': 7.21,
            'prev_hkd_rate': 0.927,
            'current_usd_rate': 7.21,
            'current_hkd_rate': 0.927,
            'year': 2025,
            'recent_files': [],
            'auto_save_log': True,
            'use_fast_mode': True
        }

    @staticmethod
    def save_config(config):
        """保存配置"""
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except:
            pass

    @staticmethod
    def add_recent_file(file_path, config):
        """添加最近文件"""
        if 'recent_files' not in config:
            config['recent_files'] = []
        if file_path in config['recent_files']:
            config['recent_files'].remove(file_path)
        config['recent_files'].insert(0, file_path)
        config['recent_files'] = config['recent_files'][:10]  # 只保留10个


class QuarterProcessWorker(QThread):
    """季度处理后台线程 - 优化版"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str, str)
    log = pyqtSignal(str)

    def __init__(self, current_file, prev_file, output_file, quarter, year,
                 current_usd_rate, current_hkd_rate, prev_usd_rate, prev_hkd_rate, fast_mode=True):
        super().__init__()
        self.current_file = current_file
        self.prev_file = prev_file
        self.output_file = output_file
        self.quarter = quarter  # 当季度
        self.year = year
        # 当季度汇率
        self.current_usd_rate = current_usd_rate
        self.current_hkd_rate = current_hkd_rate
        # 上季度汇率
        self.prev_usd_rate = prev_usd_rate
        self.prev_hkd_rate = prev_hkd_rate
        self.fast_mode = fast_mode

        # 计算上季度
        if self.quarter == 1:
            self.prev_quarter = 4
            self.prev_year = self.year - 1
        else:
            self.prev_quarter = self.quarter - 1
            self.prev_year = self.year

    def run(self):
        try:
            self.process_files()
        except Exception as e:
            import traceback
            self.finished.emit(False, str(e), traceback.format_exc())

    def parse_date_vectorized(self, series):
        """向量化日期解析 - 性能优化"""
        return pd.to_datetime(series, errors='coerce')

    def process_files(self):
        # ===== Step A: 读取文件并处理结清记录 =====
        self.progress.emit(5, "读取当季度文件...")
        self.log.emit("📂 Step A: 读取当季度文件...")
        df_current = pd.read_excel(self.current_file)
        self.log.emit(f"   当季度数据: {len(df_current)} 行")

        self.progress.emit(10, "读取上季度文件...")
        self.log.emit("📂 读取上季度文件...")
        df_prev_raw = pd.read_excel(self.prev_file, header=None)
        header_row = 0
        for idx in range(min(10, len(df_prev_raw))):
            if any('贷款编号' in str(val) for val in df_prev_raw.iloc[idx].astype(str).tolist()):
                header_row = idx
                break
        df_prev = pd.read_excel(self.prev_file, header=header_row)
        self.log.emit(f"   上季度数据: {len(df_prev)} 行")

        # 统一列名
        self.progress.emit(15, "统一列名...")
        self.log.emit("🔧 统一列名...")
        standard_cols = ['贷款编号', '经办机构', '被保险人', '被保险财产标的', '城市',
                        '被保险财产地址', '币种', '保险金额', '保险起期', '保险止期', '贷款分类']

        # 【修复v2】按列名选择标准列,确保两个文件选择相同的列
        # 找出每个标准列在原文件中的位置
        def find_standard_columns(df, standard_cols):
            """智能查找标准列,支持模糊匹配"""
            selected_cols = {}

            for std_col in standard_cols:
                # 精确匹配
                if std_col in df.columns:
                    selected_cols[std_col] = std_col
                else:
                    # 模糊匹配: 查找包含关键字的列
                    for col in df.columns:
                        if std_col in str(col):
                            selected_cols[std_col] = col
                            break

            return selected_cols

        # 查找上季度文件的标准列
        prev_col_map = find_standard_columns(df_prev, standard_cols)
        self.log.emit(f"   上季度列映射: {prev_col_map}")

        # 查找当季度文件的标准列
        curr_col_map = find_standard_columns(df_current, standard_cols)
        self.log.emit(f"   当季度列映射: {curr_col_map}")

        # 确保所有标准列都找到了
        missing_prev = [col for col in standard_cols if col not in prev_col_map]
        missing_curr = [col for col in standard_cols if col not in curr_col_map]

        if missing_prev:
            self.log.emit(f"   ⚠️ 上季度文件缺少列: {missing_prev}")
        if missing_curr:
            self.log.emit(f"   ⚠️ 当季度文件缺少列: {missing_curr}")

        # 按映射选择列并重命名为标准列名
        df_prev_clean = df_prev[[prev_col_map[col] for col in standard_cols if col in prev_col_map]].copy()
        df_prev_clean.columns = [col for col in standard_cols if col in prev_col_map]

        df_current_clean = df_current[[curr_col_map[col] for col in standard_cols if col in curr_col_map]].copy()
        df_current_clean.columns = [col for col in standard_cols if col in curr_col_map]

        self.log.emit(f"   上季度选择列: {list(df_prev_clean.columns)}")
        self.log.emit(f"   当季度选择列: {list(df_current_clean.columns)}")

        # A. 从当季度表格中剪切结清记录，粘贴到上季度末尾
        self.progress.emit(20, "处理结清记录...")
        self.log.emit("🔍 Step A: 筛选'结清'记录并移动...")

        # 计算当季度范围
        bounds = {1: (datetime(self.year,1,1), datetime(self.year,3,31)),
                  2: (datetime(self.year,4,1), datetime(self.year,6,30)),
                  3: (datetime(self.year,7,1), datetime(self.year,9,30)),
                  4: (datetime(self.year,10,1), datetime(self.year,12,31))}
        q_start, q_end = bounds[self.quarter]

        loan_class_col = '贷款分类'

        # 解析保险止期
        end_dates = pd.to_datetime(df_current_clean['保险止期'], errors='coerce')

        # 【诊断】检查保险止期解析情况
        self.log.emit(f"   [诊断] 保险止期列是否存在: {'保险止期' in df_current_clean.columns}")
        self.log.emit(f"   [诊断] 保险止期解析成功: {pd.notna(end_dates).sum()} 条")
        self.log.emit(f"   [诊断] 当季度范围: {q_start.strftime('%Y/%m/%d')} - {q_end.strftime('%Y/%m/%d')}")

        # 【关键规则v4.1.4】只移动保险止期在当季度之前的结清记录
        # 条件1: 贷款分类包含"结清"
        is_settled_class = df_current_clean[loan_class_col].astype(str).str.contains('结清', na=False)
        self.log.emit(f"   [诊断] 结清记录总数: {is_settled_class.sum()} 条")

        # 条件2: 保险止期在当季度开始之前
        end_before_quarter = end_dates < q_start
        self.log.emit(f"   [诊断] 保险止期在当季度之前: {end_before_quarter.sum()} 条")

        # 显示几个结清记录的保险止期示例
        if is_settled_class.sum() > 0:
            self.log.emit(f"   [诊断] 结清记录保险止期示例:")
            settled_sample = df_current_clean[is_settled_class].head(5)
            for idx in settled_sample.index:
                end_date = end_dates[idx]
                end_str = end_date.strftime('%Y/%m/%d') if pd.notna(end_date) else 'N/A'
                before_q = end_dates[idx] < q_start if pd.notna(end_dates[idx]) else False
                self.log.emit(f"      贷款[{settled_sample.loc[idx, '贷款编号']}]: 止期={end_str}, 早于{q_start.strftime('%Y/%m/%d')}? {before_q}")

        # 同时满足两个条件才移动
        is_settled_to_move = is_settled_class & end_before_quarter

        # 需要移动到上季度的结清记录
        settled = df_current_clean[is_settled_to_move].copy()

        # 留在当季度的记录（包括保险止期在当季度内的结清记录）
        df_remainder = df_current_clean[~is_settled_to_move].copy()

        # 统计结清记录的分类
        settled_in_quarter = is_settled_class & (~end_before_quarter)
        settled_in_quarter_count = settled_in_quarter.sum()

        self.log.emit(f"   当季度结清记录分析:")
        self.log.emit(f"   - 总结清记录: {is_settled_class.sum()} 条")
        self.log.emit(f"   - 保险止期在当季度之前: {len(settled)} 条 → 移动到上季度")
        self.log.emit(f"   - 保险止期在当季度内: {settled_in_quarter_count} 条 → 保留在当季度")
        self.log.emit(f"   当季度剩余: {len(df_remainder)} 条")

        # 输出保留在当季度的结清记录示例
        if settled_in_quarter_count > 0:
            self.log.emit(f"   📌 保留在当季度的结清记录示例:")
            kept_settled = df_current_clean[settled_in_quarter].head(3)
            for idx, row in kept_settled.iterrows():
                end_date = end_dates[idx]
                end_str = end_date.strftime('%Y/%m/%d') if pd.notna(end_date) else 'N/A'
                self.log.emit(f"      ✓ 贷款[{row['贷款编号']}]: 保险止期={end_str} (在{self.year}Q{self.quarter}内)")
            if settled_in_quarter_count > 3:
                self.log.emit(f"      ... 还有 {settled_in_quarter_count - 3} 条")

        # 粘贴到上季度末尾
        self.log.emit(f"   [诊断] 上季度列: {list(df_prev_clean.columns)}")
        self.log.emit(f"   [诊断] 结清记录列: {list(settled.columns)}")
        self.log.emit(f"   [诊断] 上季度shape: {df_prev_clean.shape}, 结清shape: {settled.shape}")
        df_prev_merged = pd.concat([df_prev_clean, settled], ignore_index=True)
        self.log.emit(f"   合并到上季度后: {len(df_prev_merged)} 行")
        self.log.emit(f"   [诊断] 合并后列: {list(df_prev_merged.columns)}")

        # A2. 删除上季度表格中贷款编号重复且含'存量'的行
        # 【修改v4.1.4】简化逻辑: 移入上季度的结清记录都是保险止期在当季度之前的
        # 可以直接删除对应的存量记录
        self.progress.emit(25, "删除重复存量...")
        self.log.emit("🔄 Step A2: 删除上季度重复存量行...")

        # 标记重复贷款编号
        df_prev_merged['_dup'] = df_prev_merged.duplicated(subset=['贷款编号'], keep=False)

        # 标记是否为存量
        df_prev_merged['_is_stock'] = df_prev_merged['贷款分类'].astype(str).str.contains('存量', na=False)

        # 删除重复贷款编号中的存量记录
        mask = df_prev_merged['_dup'] & df_prev_merged['_is_stock']
        removed = mask.sum()

        df_prev_merged = df_prev_merged[~mask].drop(columns=['_dup', '_is_stock'])

        self.log.emit(f"   删除重复存量: {removed} 条")
        self.log.emit(f"   上季度最终: {len(df_prev_merged)} 行")

        # ===== Step B: 计算人民币保险金额（分别使用不同汇率） =====
        self.progress.emit(30, "处理当季度数据...")
        self.log.emit(f"💰 Step B: 计算人民币保险金额...")
        self.log.emit(f"   上季度汇率: USD={self.prev_usd_rate}, HKD={self.prev_hkd_rate}")
        self.log.emit(f"   当季度汇率: USD={self.current_usd_rate}, HKD={self.current_hkd_rate}")

        # B1. 处理当季度剩余数据（添加人民币金额、季度日期列 - 暂不填充，Excel公式处理）
        df_remainder['人民币保险金额'] = 0.0
        df_remainder['季度起期'] = pd.NaT
        df_remainder['季度止期'] = pd.NaT

        # B2. 处理上季度数据（添加人民币金额列 - 暂不填充，Excel公式处理）
        df_prev_merged['人民币保险金额'] = 0.0

        # ===== Step B3: 检测和修正保险起期年份错误 =====
        self.progress.emit(40, "检测日期年份错误...")
        self.log.emit("🔍 Step B3: 检测和修正保险起期年份错误...")

        def detect_and_fix_year_errors(df, expected_year, quarter):
            """检测并修正保险起期年份错误

            Args:
                df: 数据框
                expected_year: 预期年份
                quarter: 当前季度(1-4)
            """
            df_copy = df.copy()

            # 解析保险起期和止期
            start_dates = pd.to_datetime(df_copy['保险起期'], errors='coerce')
            end_dates = pd.to_datetime(df_copy['保险止期'], errors='coerce')

            # 检测年份错误的条件:
            # 1. 保险起期年份与预期年份相差超过2年
            # 2. 保险止期年份在合理范围内(expected_year-1 到 expected_year+1)
            # 3. 保险起期晚于保险止期(明显错误)

            errors_found = 0
            corrected_dates = []

            for idx, (start, end) in enumerate(zip(start_dates, end_dates)):
                if pd.isna(start) or pd.isna(end):
                    continue

                start_year = start.year
                end_year = end.year

                # 条件1: 起期年份与预期相差过大(超过2年)
                year_diff_large = abs(start_year - expected_year) > 2

                # 条件2: 止期年份在合理范围
                end_year_reasonable = (expected_year - 1) <= end_year <= (expected_year + 1)

                # 条件3: 起期晚于止期
                start_after_end = start > end

                if (year_diff_large and end_year_reasonable) or start_after_end:
                    # 尝试修正: 使用止期的年份作为参考
                    # 如果保险跨年，起期可能是上一年
                    corrected_year = end_year if end.month >= 6 else end_year - 1

                    # 保持月份和日期不变
                    try:
                        corrected_start = start.replace(year=corrected_year)

                        # 验证修正后的日期是否合理(起期应早于止期)
                        if corrected_start <= end:
                            loan_no = df_copy.iloc[idx]['贷款编号']
                            original_str = start.strftime('%Y/%m/%d') if pd.notna(start) else 'N/A'
                            corrected_str = corrected_start.strftime('%Y/%m/%d')

                            corrected_dates.append({
                                'index': idx,
                                'loan_no': loan_no,
                                'original': start,
                                'corrected': corrected_start,
                                'original_str': original_str,
                                'corrected_str': corrected_str,
                                'end_date': end
                            })

                            # 修正日期
                            df_copy.at[df_copy.index[idx], '保险起期'] = corrected_start
                            errors_found += 1
                    except:
                        pass

            return df_copy, corrected_dates, errors_found

        # 检测并修正上季度数据
        df_prev_merged, prev_corrections, prev_errors = detect_and_fix_year_errors(
            df_prev_merged, self.year, self.quarter
        )

        if prev_errors > 0:
            self.log.emit(f"   🚨 上季度发现 {prev_errors} 条保险起期年份错误,已自动修正:")
            for correction in prev_corrections[:5]:  # 显示前5条
                self.log.emit(f"      ⚠️ 贷款[{correction['loan_no']}]: "
                            f"{correction['original_str']} → {correction['corrected_str']} "
                            f"(保险止期: {correction['end_date'].strftime('%Y/%m/%d')})")
            if prev_errors > 5:
                self.log.emit(f"      ... 还有 {prev_errors - 5} 条已修正")
        else:
            self.log.emit(f"   ✓ 上季度未发现年份错误")

        # 检测并修正当季度数据
        df_remainder, curr_corrections, curr_errors = detect_and_fix_year_errors(
            df_remainder, self.year, self.quarter
        )

        if curr_errors > 0:
            self.log.emit(f"   🚨 当季度发现 {curr_errors} 条保险起期年份错误,已自动修正:")
            for correction in curr_corrections[:5]:  # 显示前5条
                self.log.emit(f"      ⚠️ 贷款[{correction['loan_no']}]: "
                            f"{correction['original_str']} → {correction['corrected_str']} "
                            f"(保险止期: {correction['end_date'].strftime('%Y/%m/%d')})")
            if curr_errors > 5:
                self.log.emit(f"      ... 还有 {curr_errors - 5} 条已修正")
        else:
            self.log.emit(f"   ✓ 当季度未发现年份错误")

        # ===== Step C: 计算季度起期和季度止期 =====
        self.progress.emit(50, "计算季度日期...")
        self.log.emit(f"📅 Step C: 计算季度日期 (上季度: {self.prev_year}Q{self.prev_quarter}, 当季度: {self.year}Q{self.quarter})...")

        # 计算当季度范围(用于筛选结清记录)
        bounds = {1: (datetime(self.year,1,1), datetime(self.year,3,31)),
                  2: (datetime(self.year,4,1), datetime(self.year,6,30)),
                  3: (datetime(self.year,7,1), datetime(self.year,9,30)),
                  4: (datetime(self.year,10,1), datetime(self.year,12,31))}
        q_start, q_end = bounds[self.quarter]

        # 计算上季度范围(用于计算上季度文件的季度日期)
        prev_bounds = {1: (datetime(self.prev_year,1,1), datetime(self.prev_year,3,31)),
                       2: (datetime(self.prev_year,4,1), datetime(self.prev_year,6,30)),
                       3: (datetime(self.prev_year,7,1), datetime(self.prev_year,9,30)),
                       4: (datetime(self.prev_year,10,1), datetime(self.prev_year,12,31))}
        prev_q_start, prev_q_end = prev_bounds[self.prev_quarter]

        def calc_quarter_dates(df, q_start, q_end, apply_settled_rule=False):
            """计算季度起期和止期

            Args:
                df: 数据框
                q_start: 季度开始日期
                q_end: 季度结束日期
                apply_settled_rule: 是否应用结清记录特殊规则
            """
            df_copy = df.copy()

            # C. 向量化日期解析
            start_dates = self.parse_date_vectorized(df_copy['保险起期'])
            end_dates = self.parse_date_vectorized(df_copy['保险止期'])

            # 标记是否为结清记录
            is_settled = df_copy['贷款分类'].astype(str).str.contains('结清', na=False)

            # 标记保险止期是否在当季度内
            end_in_quarter = (end_dates >= q_start) & (end_dates <= q_end)

            # C. 计算季度起期
            df_copy['季度起期'] = start_dates
            df_copy.loc[start_dates < q_start, '季度起期'] = q_start
            df_copy.loc[start_dates > q_end, '季度起期'] = pd.NaT

            # 【修正v4.1.6】不再强制结清记录的季度起期为季度第一天
            # 结清记录遵循标准规则:
            # - 如果保险起期 >= 季度开始, 季度起期 = 保险起期
            # - 如果保险起期 < 季度开始, 季度起期 = 季度开始

            # C. 计算季度止期
            df_copy['季度止期'] = end_dates
            df_copy.loc[end_dates > q_end, '季度止期'] = q_end
            df_copy.loc[end_dates < q_start, '季度止期'] = pd.NaT

            # 【修正v4.1.6】结清记录的季度止期使用实际保险止期(不截断到季度末)
            # 这样可以准确计算当季度的保费天数
            if apply_settled_rule:
                settled_in_q = is_settled & end_in_quarter
                if settled_in_q.any():
                    df_copy.loc[settled_in_q, '季度止期'] = end_dates[settled_in_q]
                    # 输出示例
                    count = settled_in_q.sum()
                    self.log.emit(f"   ✓ 应用结清规则: {count} 条记录")
                    examples = df_copy[settled_in_q].head(3)
                    for idx, row in examples.iterrows():
                        q_start_actual = row['季度起期']
                        q_end_actual = row['季度止期']
                        self.log.emit(f"      结清[{row['贷款编号']}]: 季度起期={q_start_actual.strftime('%Y/%m/%d') if pd.notna(q_start_actual) else 'N/A'}, 季度止期={q_end_actual.strftime('%Y/%m/%d') if pd.notna(q_end_actual) else 'N/A'}")

            # 判断是否在季度内
            df_copy['_in_q'] = (~start_dates.isna()) & (~end_dates.isna()) & \
                                 ~((end_dates < q_start) | (start_dates > q_end))

            return df_copy

        # C1. 计算上季度的季度日期(使用上季度范围,应用结清规则)
        df_prev_merged = calc_quarter_dates(df_prev_merged, prev_q_start, prev_q_end, apply_settled_rule=True)
        self.log.emit(f"   上季度({self.prev_year}Q{self.prev_quarter})季度日期计算完成")

        # C2. 计算当季度的季度日期(使用当季度范围,应用结清规则,因为可能包含保险止期在当季度内的结清记录)
        df_remainder = calc_quarter_dates(df_remainder, q_start, q_end, apply_settled_rule=True)
        self.log.emit(f"   当季度({self.year}Q{self.quarter})季度日期计算完成")

        # ===== Step D: 添加季度保费列（暂不填充，Excel公式处理） =====
        self.progress.emit(60, "准备季度保费列...")
        self.log.emit("💵 Step D: 准备季度保费列...")
        df_prev_merged['季度保费'] = 0.0

        # ===== Step D2: 最终清理 - 删除重复贷款编号中的存量记录 =====
        self.progress.emit(65, "最终清理重复记录...")
        self.log.emit("🧹 Step D2: 最终清理 - 删除重复贷款编号中的存量记录...")

        # 统计清理前的记录数
        before_cleanup = len(df_prev_merged)

        # 标记重复的贷款编号
        df_prev_merged['_is_dup'] = df_prev_merged.duplicated(subset=['贷款编号'], keep=False)

        # 标记是否为存量
        df_prev_merged['_is_stock'] = df_prev_merged['贷款分类'].astype(str).str.contains('存量', na=False)

        # 找出所有重复的贷款编号
        dup_loans = df_prev_merged[df_prev_merged['_is_dup']]['贷款编号'].unique()

        # 删除规则: 贷款编号重复 且 贷款分类为存量
        to_remove_final = []
        for loan_no in dup_loans:
            loan_records = df_prev_merged[df_prev_merged['贷款编号'] == loan_no]
            stock_records = loan_records[loan_records['_is_stock']]

            if len(stock_records) > 0:
                # 记录要删除的存量记录
                to_remove_final.extend(stock_records.index.tolist())
                self.log.emit(f"   🗑️ 删除重复贷款[{loan_no}]的存量记录({len(stock_records)}条)")

        # 执行删除
        df_prev_merged = df_prev_merged.drop(index=to_remove_final)
        df_prev_merged = df_prev_merged.drop(columns=['_is_dup', '_is_stock'])

        removed_final = len(to_remove_final)
        after_cleanup = len(df_prev_merged)

        self.log.emit(f"   最终清理完成:")
        self.log.emit(f"   - 清理前: {before_cleanup} 行")
        self.log.emit(f"   - 删除存量: {removed_final} 条")
        self.log.emit(f"   - 清理后: {after_cleanup} 行")

        # ===== Step E: 生成Excel（使用公式计算人民币金额） =====
        self.progress.emit(70, "生成Excel...")
        self.log.emit("📊 Step E: 生成Excel输出...")

        # E1. 先保存当季度剔除结清后的文件（带公式）
        self.log.emit("💾 保存当季度剔除结清后文件...")
        remainder_file = str(Path(self.output_file).parent / f"{Path(self.current_file).stem}_剔除结清后.xlsx")

        # 准备当季度输出列
        current_output_cols = ['贷款编号', '经办机构', '被保险人', '被保险财产标的', '城市', '被保险财产地址',
                              '币种', '保险金额', '人民币保险金额', '保险起期', '保险止期', '贷款分类',
                              '季度起期', '季度止期']
        df_current_output = df_remainder[current_output_cols].copy()

        with pd.ExcelWriter(remainder_file, engine='openpyxl') as writer:
            df_current_output.to_excel(writer, sheet_name=f"{self.year}Q{self.quarter}", index=False)

        # E2. 准备上季度输出数据
        output_cols = ['贷款编号', '经办机构', '被保险人', '被保险财产标的', '城市', '被保险财产地址',
                       '币种', '保险金额', '人民币保险金额', '保险起期', '保险止期', '贷款分类',
                       '季度起期', '季度止期', '季度保费']
        df_output = df_prev_merged[output_cols].copy()

        # 使用 ExcelWriter 批量写入 (使用上季度的年份和季度)
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df_output.to_excel(writer, sheet_name=f"{self.prev_year}Q{self.prev_quarter}", index=False)

            # 写入空的分行统计（稍后用公式计算）
            stats_headers = ['经办机构', '季度保费合计', '人民币保险金额合计', '贷款笔数']
            df_stats_empty = pd.DataFrame(columns=stats_headers)
            df_stats_empty.to_excel(writer, sheet_name='分行统计', index=False)

        # E3. 格式化当季度文件并添加公式
        self.log.emit("🎨 格式化当季度文件...")
        wb_current = load_workbook(remainder_file)
        ws_current = wb_current[f"{self.year}Q{self.quarter}"]

        cfont = Font(name='Microsoft YaHei', size=9)
        border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                       top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

        # 添加人民币金额公式（当季度汇率）
        self.log.emit(f"   添加当季度人民币金额公式 (USD={self.current_usd_rate}, HKD={self.current_hkd_rate})...")
        for row_idx in range(2, len(df_remainder) + 2):
            # I列 (col=9): 人民币保险金额公式
            formula = f'=H{row_idx}*IF(G{row_idx}="美元",{self.current_usd_rate},IF(G{row_idx}="港币",{self.current_hkd_rate},1))'
            ws_current.cell(row=row_idx, column=9, value=formula)

        # 设置格式
        for row_idx in range(1, len(df_remainder) + 2):
            for col_idx in range(1, 15):
                cell = ws_current.cell(row=row_idx, column=col_idx)
                cell.font = cfont
                cell.border = border
                if row_idx >= 2 and col_idx <= 6:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(col_idx == 6))
                elif row_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 6))

        # 设置数字格式
        for row in ws_current.iter_rows(min_row=2, max_row=len(df_remainder) + 1, min_col=8, max_col=9):
            for cell in row:
                cell.number_format = '#,##0.00'

        # 设置日期格式
        date_cols = [10, 11, 13, 14]
        for row in ws_current.iter_rows(min_row=2, max_row=len(df_remainder) + 1):
            for col_idx in date_cols:
                cell = row[col_idx - 1]
                if cell.value:
                    cell.number_format = 'YYYY/M/D'

        # 设置列宽
        widths = {'A': 20, 'B': 10, 'C': 12, 'D': 14, 'E': 8, 'F': 56, 'G': 6, 'H': 14, 'I': 16,
                  'J': 12, 'K': 12, 'L': 8, 'M': 12, 'N': 12}
        for col, w in widths.items():
            ws_current.column_dimensions[col].width = w

        wb_current.save(remainder_file)
        self.log.emit(f"   ✅ 当季度文件保存完成: {len(df_remainder)} 行")

        # E4. 格式化上季度文件并添加公式
        self.progress.emit(80, "格式化上季度Excel...")
        self.log.emit("🎨 格式化上季度文件...")
        wb = load_workbook(self.output_file)
        ws = wb[f"{self.prev_year}Q{self.prev_quarter}"]

        # 添加人民币金额公式（上季度汇率）
        self.log.emit(f"   添加上季度人民币金额公式 (USD={self.prev_usd_rate}, HKD={self.prev_hkd_rate})...")
        for row_idx in range(2, len(df_prev_merged) + 2):
            # I列 (col=9): 人民币保险金额公式
            formula = f'=H{row_idx}*IF(G{row_idx}="美元",{self.prev_usd_rate},IF(G{row_idx}="港币",{self.prev_hkd_rate},1))'
            ws.cell(row=row_idx, column=9, value=formula)

            # O列 (col=15): 季度保费公式
            # =(N-M+1)*I/365*0.0037/100
            formula_premium = f'=IF(AND(M{row_idx}<>"",N{row_idx}<>""),(N{row_idx}-M{row_idx}+1)*I{row_idx}/365*0.0037/100,0)'
            ws.cell(row=row_idx, column=15, value=formula_premium)

        # 【不改变填充色】只设置字体、边框和对齐方式
        for row_idx in range(1, len(df_prev_merged) + 2):
            for col_idx in range(1, 16):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cfont
                cell.border = border

                # A-F列左对齐（从第2行开始）
                if row_idx >= 2 and col_idx <= 6:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(col_idx == 6))
                elif row_idx == 1:
                    # 表头居中
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 6))

        # 设置数字格式（保险金额、人民币保险金额、季度保费）
        for row in ws.iter_rows(min_row=2, max_row=len(df_prev_merged) + 1, min_col=8, max_col=9):
            for cell in row:
                cell.number_format = '#,##0.00'
        for row in ws.iter_rows(min_row=2, max_row=len(df_prev_merged) + 1, min_col=15, max_col=15):
            for cell in row:
                cell.number_format = '#,##0.00'

        # 【重要】设置日期格式为 '2025/1/1' 格式
        # 保险起期 (J列=10)、保险止期 (K列=11)、季度起期 (M列=13)、季度止期 (N列=14)
        date_cols = [10, 11, 13, 14]
        for row in ws.iter_rows(min_row=2, max_row=len(df_prev_merged) + 1):
            for col_idx in date_cols:
                cell = row[col_idx - 1]  # 索引从0开始
                if cell.value:
                    cell.number_format = 'YYYY/M/D'

        # 添加汇总信息和汇率
        headers = ['贷款编号', '经办机构', '被保险人', '被保险财产标的', '城市', '被保险财产地址',
                   '币种', '保险金额', '人民币保险金额', '保险起期', '保险止期', '贷款分类',
                   '季度起期', '季度止期', '季度保费', '人民币保险金额合计', '季度保费合计', '上季度美元汇率', '上季度港币汇率']
        for c, h in enumerate(headers[15:], 16):
            ws.cell(row=1, column=c, value=h).font = cfont
            ws.cell(row=1, column=c).border = border
            ws.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center')

        last = len(df_prev_merged) + 1
        ws.cell(row=2, column=16, value=f'=SUM(I2:I{last})').number_format = '#,##0.00'
        ws.cell(row=2, column=17, value=f'=SUM(O2:O{last})').number_format = '#,##0.00'
        ws.cell(row=2, column=18, value=self.prev_usd_rate).number_format = '0.0000'
        ws.cell(row=2, column=19, value=self.prev_hkd_rate).number_format = '0.0000'
        for c in [16, 17, 18, 19]:
            ws.cell(row=2, column=c).font = Font(name='Microsoft YaHei', size=9, bold=True, color='0066CC' if c >= 18 else '000000')
            ws.cell(row=2, column=c).border = border

        # 设置列宽（F列被保险财产地址宽度调整为56）
        widths = {'A': 20, 'B': 10, 'C': 12, 'D': 14, 'E': 8, 'F': 56, 'G': 6, 'H': 14, 'I': 16,
                  'J': 12, 'K': 12, 'L': 8, 'M': 12, 'N': 12, 'O': 12, 'P': 18, 'Q': 14, 'R': 10, 'S': 10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

        self.progress.emit(90, "生成分行统计...")
        self.log.emit("📈 生成分行统计...")

        # 生成分行统计数据（通过重新读取带公式的数据）
        # 由于使用了公式，我们需要从源数据生成统计
        wb.save(self.output_file)

        # 重新加载以计算公式值
        wb = load_workbook(self.output_file, data_only=False)
        ws = wb[f"{self.prev_year}Q{self.prev_quarter}"]

        # 获取不同的经办机构
        agencies = df_prev_merged['经办机构'].unique()

        # 美化分行统计表
        ws2 = wb['分行统计']

        # 填充分行统计数据
        row_idx = 2
        for agency in sorted(agencies):
            ws2.cell(row=row_idx, column=1, value=agency)

            # 季度保费合计 - 使用SUMIF公式
            ws2.cell(row=row_idx, column=2, value=f'=SUMIF({ws.title}!$B$2:$B${len(df_prev_merged)+1},A{row_idx},{ws.title}!$O$2:$O${len(df_prev_merged)+1})')

            # 人民币保险金额合计 - 使用SUMIF公式
            ws2.cell(row=row_idx, column=3, value=f'=SUMIF({ws.title}!$B$2:$B${len(df_prev_merged)+1},A{row_idx},{ws.title}!$I$2:$I${len(df_prev_merged)+1})')

            # 贷款笔数 - 使用COUNTIF公式
            ws2.cell(row=row_idx, column=4, value=f'=COUNTIF({ws.title}!$B$2:$B${len(df_prev_merged)+1},A{row_idx})')

            row_idx += 1

        # 设置格式
        for col_idx in range(1, 5):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = cfont
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 格式化数据行
        num_rows = len(agencies)
        for r in range(2, num_rows + 2):
            for col_idx in range(1, 5):
                cell = ws2.cell(row=r, column=col_idx)
                cell.border = border
                cell.font = cfont
                if col_idx >= 2:
                    cell.number_format = '#,##0.00' if col_idx <= 3 else '0'

        # 合计行
        tr = num_rows + 2
        tfont = Font(name='Microsoft YaHei', size=9, bold=True)
        ws2.cell(row=tr, column=1, value='合计').font = tfont
        ws2.cell(row=tr, column=2, value=f'=SUM(B2:B{tr - 1})').font = tfont
        ws2.cell(row=tr, column=2).number_format = '#,##0.00'
        ws2.cell(row=tr, column=3, value=f'=SUM(C2:C{tr - 1})').number_format = '#,##0.00'
        ws2.cell(row=tr, column=4, value=f'=SUM(D2:D{tr - 1})').number_format = '0'
        for c in range(1, 5):
            ws2.cell(row=tr, column=c).border = border
            ws2.cell(row=tr, column=c).font = tfont

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 10
        ws2.freeze_panes = 'A2'

        wb.save(self.output_file)

        self.log.emit(f"\n{'='*50}")
        self.log.emit(f"✅ 完成! 上季度记录:{len(df_prev_merged)} 行")
        self.log.emit(f"   当季度剔除结清后:{len(df_remainder)} 行")
        self.log.emit(f"   输出文件: {Path(self.output_file).name}")
        self.log.emit(f"   当季度文件: {Path(remainder_file).name}")
        self.log.emit(f"{'='*50}")
        self.progress.emit(100, "完成!")
        self.finished.emit(True, self.output_file, "")


class YearCompareWorker(QThread):
    """年度对比后台线程"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str, str)
    log = pyqtSignal(str)

    def __init__(self, q1, q2, q3, q4, annual, output, year):
        super().__init__()
        self.q_files = {'Q1': q1, 'Q2': q2, 'Q3': q3, 'Q4': q4}
        self.annual_file = annual
        self.output_file = output
        self.year = year

    def run(self):
        try:
            self.compare()
        except Exception as e:
            import traceback
            self.finished.emit(False, str(e), traceback.format_exc())

    def read_branch(self, path, name):
        """读取分行统计"""
        self.log.emit(f"   {name}: {Path(path).name}")
        try:
            df = pd.read_excel(path, sheet_name='分行统计')
            prem_col = next((c for c in df.columns if '季度保费' in str(c)), df.columns[-1])
            branch_col = df.columns[0]
            df = df[~df[branch_col].astype(str).str.contains('合计', na=False)]
            result = df[[branch_col, prem_col]].rename(columns={branch_col: '经办机构', prem_col: name})
            # 检查数据是否为空（公式未计算的情况）
            if result[name].notna().sum() > 0:
                return result
        except:
            pass
        # 回退：从原始数据表计算季度保费
        df = pd.read_excel(path, sheet_name=0)
        branch = next((c for c in df.columns if '经办机构' in str(c)), None)
        prem_col = next((c for c in df.columns if c == '季度保费'), None)
        # 如果季度保费列有数据，直接使用
        if prem_col and df[prem_col].notna().sum() > 0:
            return df.groupby(branch)[prem_col].sum().reset_index().rename(columns={branch: '经办机构', prem_col: name})
        # 否则根据保险金额、币种、汇率计算
        if not branch: raise ValueError("找不到经办机构列")
        amt_col = next((c for c in df.columns if c == '保险金额'), None)
        curr_col = next((c for c in df.columns if c == '币种'), None)
        if not amt_col or not curr_col: raise ValueError("找不到保险金额或币种列")
        # 获取汇率（从文件中读取或使用默认值）
        usd_col = next((c for c in df.columns if '美元汇率' in str(c)), None)
        hkd_col = next((c for c in df.columns if '港币汇率' in str(c)), None)
        usd_rate = df[usd_col].dropna().iloc[0] if usd_col and df[usd_col].notna().any() else 7.2
        hkd_rate = df[hkd_col].dropna().iloc[0] if hkd_col and df[hkd_col].notna().any() else 0.93
        # 计算人民币保险金额
        def to_rmb(row):
            amt = row[amt_col] if pd.notna(row[amt_col]) else 0
            curr = row[curr_col] if pd.notna(row[curr_col]) else ''
            if curr == '美元': return amt * usd_rate
            if curr == '港币': return amt * hkd_rate
            return amt
        df['_rmb_amt'] = df.apply(to_rmb, axis=1)
        # 季度保费 = 人民币保险金额 * 0.00025，四舍五入保留两位小数
        df['_prem_raw'] = df['_rmb_amt'] * 0.00025
        # 四舍五入后，如果原值>0但结果为0，则调整为0.01
        df['_prem'] = df['_prem_raw'].apply(lambda x: 0.01 if x > 0 and round(x, 2) == 0 else round(x, 2))
        return df.groupby(branch)['_prem'].sum().reset_index().rename(columns={branch: '经办机构', '_prem': name})

    def read_annual(self, path):
        """读取年初总表"""
        self.log.emit(f"   年初总表: {Path(path).name}")
        try:
            df = pd.read_excel(path, sheet_name='分行统计')
            prem_col = next((c for c in df.columns if '保费' in str(c)), df.columns[-1])
            branch_col = df.columns[0]
            df = df[~df[branch_col].astype(str).str.contains('合计', na=False)]
            return df[[branch_col, prem_col]].rename(columns={branch_col: '经办机构', prem_col: '年初保费'})
        except:
            raw = pd.read_excel(path, header=None)
            hrow = next((i for i in range(min(10, len(raw))) 
                        if any('贷款编号' in str(v) or '经办机构' in str(v) for v in raw.iloc[i].astype(str))), 0)
            df = pd.read_excel(path, header=hrow)
            branch = next((c for c in df.columns if '经办机构' in str(c)), None)
            prem = next((c for c in df.columns if '保费' in str(c)), None)
            if not branch or not prem: raise ValueError("找不到必要列")
            return df.groupby(branch)[prem].sum().reset_index().rename(columns={branch: '经办机构', prem: '年初保费'})

    def compare(self):
        self.progress.emit(5, "开始对比...")
        self.log.emit("📊 年度对比分析\n" + "="*50)
        
        self.log.emit("\n📂 读取季度文件...")
        q_data = []
        for i, (q, path) in enumerate(self.q_files.items()):
            self.progress.emit(10 + i*15, f"读取{q}...")
            if path and os.path.exists(path):
                q_data.append(self.read_branch(path, f'{q}保费'))
                self.log.emit(f"      ✓ {len(q_data[-1])} 分行")
            else:
                self.log.emit(f"      ⚠️ {q}未选择")
        
        self.progress.emit(70, "读取年初总表...")
        self.log.emit("\n📂 读取年初总表...")
        annual = self.read_annual(self.annual_file)
        self.log.emit(f"      ✓ {len(annual)} 分行")
        
        self.progress.emit(80, "合并数据...")
        self.log.emit("\n🔄 合并对比...")
        merged = annual.copy()
        for df in q_data:
            merged = merged.merge(df, on='经办机构', how='outer')
        merged = merged.fillna(0)
        
        q_cols = [c for c in merged.columns if c.startswith('Q') and '保费' in c]
        merged['四季度累加'] = merged[q_cols].sum(axis=1)
        merged['差额'] = merged['四季度累加'] - merged['年初保费']
        merged = merged.sort_values('差额', key=lambda x: x.abs(), ascending=False)
        
        self.progress.emit(90, "生成报告...")
        self.log.emit("\n📊 生成报告...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.year}年度对比"
        
        hfont = Font(name='Microsoft YaHei', size=9, bold=True, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='2563eb')
        cfont = Font(name='Microsoft YaHei', size=9)
        pos_font = Font(name='Microsoft YaHei', size=9, color='059669')
        neg_font = Font(name='Microsoft YaHei', size=9, color='DC2626')
        border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                       top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
        
        headers = ['经办机构', '年初保费'] + q_cols + ['四季度累加', '差额']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill, cell.border = hfont, hfill, border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for i, (_, row) in enumerate(merged.iterrows(), 2):
            for c, col in enumerate(headers, 1):
                val = row.get(col, 0)
                cell = ws.cell(row=i, column=c, value=val)
                cell.border = border
                if col == '经办机构':
                    cell.font = cfont
                elif col == '差额':
                    cell.font = pos_font if val > 0 else (neg_font if val < 0 else cfont)
                    cell.number_format = '#,##0.00'
                else:
                    cell.font = cfont
                    cell.number_format = '#,##0.00'
        
        tr = len(merged) + 2
        tfill = PatternFill('solid', fgColor='E8E8E8')
        tfont = Font(name='Microsoft YaHei', size=9, bold=True)
        ws.cell(row=tr, column=1, value='合计').font = tfont
        ws.cell(row=tr, column=1).fill = tfill
        ws.cell(row=tr, column=1).border = border
        
        for c in range(2, len(headers)+1):
            cell = ws.cell(row=tr, column=c, value=f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}{tr-1})')
            cell.font = tfont
            cell.fill = tfill
            cell.number_format = '#,##0.00'
            cell.border = border
        
        diff_sum = merged['差额'].sum()
        ws.cell(row=tr, column=len(headers)).font = Font(name='Microsoft YaHei', size=9, bold=True, 
                                                          color='059669' if diff_sum >= 0 else 'DC2626')
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        for i in range(3, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 12 if i < len(headers)-1 else 14
        ws.freeze_panes = 'A2'
        
        wb.save(self.output_file)
        
        self.log.emit(f"\n{'='*50}")
        self.log.emit(f"✅ 对比完成!")
        self.log.emit(f"   年初总保费: ¥{merged['年初保费'].sum():,.2f}")
        self.log.emit(f"   四季度累加: ¥{merged['四季度累加'].sum():,.2f}")
        self.log.emit(f"   总差额: ¥{diff_sum:,.2f}")
        self.log.emit(f"{'='*50}")
        
        self.progress.emit(100, "完成!")
        self.finished.emit(True, self.output_file, "")


class InsuranceFeeProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.q_current_file = self.q_prev_file = self.q_output_file = None
        self.c_q_files = {}
        self.c_annual_file = self.c_output_file = None

        # 加载配置
        self.config = ConfigManager.load_config()

        self.init_ui()
        self.load_settings()

    def init_ui(self):
        self.setWindowTitle("保险费用Excel处理工具 v4.0 (优化版)")
        self.setMinimumSize(1000, 850)
        self.setStyleSheet(STYLE_SHEET)
        
        main = QWidget()
        self.setCentralWidget(main)
        layout = QVBoxLayout(main)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(15)

        # 标题
        title = QLabel("📊 保险费用Excel处理工具 v4.0")
        title.setObjectName("titleLabel")
        subtitle = QLabel("性能优化 | 配置保存 | 批量处理 | 智能分析")
        subtitle.setObjectName("subtitleLabel")
        layout.addWidget(title)
        layout.addWidget(subtitle)

        # Tab
        tabs = QTabWidget()
        tabs.addTab(self._create_quarter_tab(), "📊 季度处理")
        tabs.addTab(self._create_compare_tab(), "📈 年度对比")
        tabs.addTab(self._create_settings_tab(), "⚙️ 设置")
        layout.addWidget(tabs)

    def _create_quarter_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        
        # 文件
        fg = QGroupBox("📁 文件选择")
        fl = QGridLayout(fg)
        fl.addWidget(QLabel("当季度文件:"), 0, 0)
        self.q_cur_path = QLineEdit()
        self.q_cur_path.setReadOnly(True)
        self.q_cur_path.setPlaceholderText("选择当季度Excel文件")
        fl.addWidget(self.q_cur_path, 0, 1)
        b1 = QPushButton("浏览...")
        b1.clicked.connect(lambda: self._sel_q_file('cur'))
        fl.addWidget(b1, 0, 2)
        
        fl.addWidget(QLabel("上季度文件:"), 1, 0)
        self.q_prev_path = QLineEdit()
        self.q_prev_path.setReadOnly(True)
        self.q_prev_path.setPlaceholderText("选择上季度Excel文件")
        fl.addWidget(self.q_prev_path, 1, 1)
        b2 = QPushButton("浏览...")
        b2.clicked.connect(lambda: self._sel_q_file('prev'))
        fl.addWidget(b2, 1, 2)
        layout.addWidget(fg)
        
        # 参数
        pg = QGroupBox("⚙️ 参数设置")
        pl = QGridLayout(pg)

        # 年份和季度
        pl.addWidget(QLabel("年份:"), 0, 0)
        self.q_year = QDoubleSpinBox()
        self.q_year.setDecimals(0)
        self.q_year.setRange(2020, 2030)
        self.q_year.setValue(2025)
        pl.addWidget(self.q_year, 0, 1)

        pl.addWidget(QLabel("季度:"), 0, 2)
        self.q_quarter = QComboBox()
        self.q_quarter.addItems(['Q1', 'Q2', 'Q3', 'Q4'])
        self.q_quarter.setCurrentIndex(2)
        pl.addWidget(self.q_quarter, 0, 3)

        # 上季度汇率
        pl.addWidget(QLabel("上季度美元汇率:"), 1, 0)
        self.q_prev_usd = QDoubleSpinBox()
        self.q_prev_usd.setDecimals(4)
        self.q_prev_usd.setRange(1, 20)
        self.q_prev_usd.setValue(7.21)
        pl.addWidget(self.q_prev_usd, 1, 1)

        pl.addWidget(QLabel("上季度港币汇率:"), 1, 2)
        self.q_prev_hkd = QDoubleSpinBox()
        self.q_prev_hkd.setDecimals(4)
        self.q_prev_hkd.setRange(0.1, 5)
        self.q_prev_hkd.setValue(0.927)
        pl.addWidget(self.q_prev_hkd, 1, 3)

        # 当季度汇率
        pl.addWidget(QLabel("当季度美元汇率:"), 2, 0)
        self.q_current_usd = QDoubleSpinBox()
        self.q_current_usd.setDecimals(4)
        self.q_current_usd.setRange(1, 20)
        self.q_current_usd.setValue(7.21)
        pl.addWidget(self.q_current_usd, 2, 1)

        pl.addWidget(QLabel("当季度港币汇率:"), 2, 2)
        self.q_current_hkd = QDoubleSpinBox()
        self.q_current_hkd.setDecimals(4)
        self.q_current_hkd.setRange(0.1, 5)
        self.q_current_hkd.setValue(0.927)
        pl.addWidget(self.q_current_hkd, 2, 3)

        layout.addWidget(pg)
        
        # 日志
        lg = QGroupBox("📋 处理日志")
        ll = QVBoxLayout(lg)
        self.q_log = QTextEdit()
        self.q_log.setReadOnly(True)
        self.q_log.setMinimumHeight(180)
        ll.addWidget(self.q_log)
        layout.addWidget(lg)
        
        # 底部
        self.q_prog = QProgressBar()
        self.q_status = QLabel("就绪")
        self.q_status.setMinimumWidth(100)
        pl2 = QHBoxLayout()
        pl2.addWidget(self.q_prog)
        pl2.addWidget(self.q_status)
        layout.addLayout(pl2)
        
        bl = QHBoxLayout()
        bl.addStretch()
        self.q_btn = QPushButton("🚀 开始处理")
        self.q_btn.setObjectName("successBtn")
        self.q_btn.clicked.connect(self._start_quarter)
        bl.addWidget(self.q_btn)
        self.q_open = QPushButton("📂 打开文件")
        self.q_open.clicked.connect(lambda: self._open('q'))
        self.q_open.setEnabled(False)
        bl.addWidget(self.q_open)
        layout.addLayout(bl)
        
        return tab

    def _create_compare_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        
        # 季度文件
        qg = QGroupBox("📁 季度处理结果文件")
        ql = QGridLayout(qg)
        ql.setColumnStretch(1, 1)
        ql.setSpacing(10)  # 行间距
        self.c_paths = {}
        for i, q in enumerate(['Q1', 'Q2', 'Q3', 'Q4']):
            lbl = QLabel(f"{q}文件:")
            lbl.setMinimumWidth(70)
            ql.addWidget(lbl, i, 0)
            p = QLineEdit()
            p.setReadOnly(True)
            p.setPlaceholderText(f"选择{q}处理结果")
            p.setMinimumSize(550, 28)  # 宽550，高38
            self.c_paths[q] = p
            ql.addWidget(p, i, 1)
            b = QPushButton("浏览...")
            b.setMinimumSize(100, 28)  # 宽100，高38
            b.clicked.connect(lambda _, x=q: self._sel_c_file(x))
            ql.addWidget(b, i, 2)
        layout.addWidget(qg)
        
        # 年初总表
        ag = QGroupBox("📁 年初总表")
        al = QGridLayout(ag)
        al.addWidget(QLabel("年初总表:"), 0, 0)
        self.c_annual_path = QLineEdit()
        self.c_annual_path.setReadOnly(True)
        self.c_annual_path.setPlaceholderText("选择年初总表")
        al.addWidget(self.c_annual_path, 0, 1)
        ab = QPushButton("浏览...")
        ab.clicked.connect(lambda: self._sel_c_file('annual'))
        al.addWidget(ab, 0, 2)
        
        al.addWidget(QLabel("年份:"), 1, 0)
        self.c_year = QDoubleSpinBox()
        self.c_year.setDecimals(0)
        self.c_year.setRange(2020, 2030)
        self.c_year.setValue(2025)
        al.addWidget(self.c_year, 1, 1)
        layout.addWidget(ag)
        
        # 日志
        lg = QGroupBox("📋 对比日志")
        ll = QVBoxLayout(lg)
        self.c_log = QTextEdit()
        self.c_log.setReadOnly(True)
        self.c_log.setMinimumHeight(180)
        ll.addWidget(self.c_log)
        layout.addWidget(lg)
        
        # 底部
        self.c_prog = QProgressBar()
        self.c_status = QLabel("就绪")
        self.c_status.setMinimumWidth(100)
        pl = QHBoxLayout()
        pl.addWidget(self.c_prog)
        pl.addWidget(self.c_status)
        layout.addLayout(pl)
        
        bl = QHBoxLayout()
        bl.addStretch()
        self.c_btn = QPushButton("🔍 开始对比")
        self.c_btn.setObjectName("compareBtn")
        self.c_btn.clicked.connect(self._start_compare)
        bl.addWidget(self.c_btn)
        self.c_open = QPushButton("📂 打开文件")
        self.c_open.clicked.connect(lambda: self._open('c'))
        self.c_open.setEnabled(False)
        bl.addWidget(self.c_open)
        layout.addLayout(bl)
        
        return tab

    def _sel_q_file(self, t):
        f, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx *.xls)")
        if f:
            if t == 'cur':
                self.q_current_file = f
                self.q_cur_path.setText(f)
            else:
                self.q_prev_file = f
                self.q_prev_path.setText(f)
            self._qlog(f"已选择: {Path(f).name}")

    def _sel_c_file(self, t):
        f, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx *.xls)")
        if f:
            if t == 'annual':
                self.c_annual_file = f
                self.c_annual_path.setText(f)
            else:
                self.c_q_files[t] = f
                self.c_paths[t].setText(f)
            self._clog(f"已选择 {t}: {Path(f).name}")

    def _create_settings_tab(self):
        """创建设置页面"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)

        # 默认汇率设置
        rate_group = QGroupBox("💱 默认汇率设置")
        rate_layout = QGridLayout(rate_group)

        # 上季度汇率
        rate_layout.addWidget(QLabel("上季度美元汇率:"), 0, 0)
        self.settings_prev_usd = QDoubleSpinBox()
        self.settings_prev_usd.setDecimals(4)
        self.settings_prev_usd.setRange(1, 20)
        self.settings_prev_usd.setValue(self.config.get('prev_usd_rate', 7.21))
        rate_layout.addWidget(self.settings_prev_usd, 0, 1)

        rate_layout.addWidget(QLabel("上季度港币汇率:"), 0, 2)
        self.settings_prev_hkd = QDoubleSpinBox()
        self.settings_prev_hkd.setDecimals(4)
        self.settings_prev_hkd.setRange(0.1, 5)
        self.settings_prev_hkd.setValue(self.config.get('prev_hkd_rate', 0.927))
        rate_layout.addWidget(self.settings_prev_hkd, 0, 3)

        # 当季度汇率
        rate_layout.addWidget(QLabel("当季度美元汇率:"), 1, 0)
        self.settings_current_usd = QDoubleSpinBox()
        self.settings_current_usd.setDecimals(4)
        self.settings_current_usd.setRange(1, 20)
        self.settings_current_usd.setValue(self.config.get('current_usd_rate', 7.21))
        rate_layout.addWidget(self.settings_current_usd, 1, 1)

        rate_layout.addWidget(QLabel("当季度港币汇率:"), 1, 2)
        self.settings_current_hkd = QDoubleSpinBox()
        self.settings_current_hkd.setDecimals(4)
        self.settings_current_hkd.setRange(0.1, 5)
        self.settings_current_hkd.setValue(self.config.get('current_hkd_rate', 0.927))
        rate_layout.addWidget(self.settings_current_hkd, 1, 3)

        layout.addWidget(rate_group)

        # 其他设置
        other_group = QGroupBox("⚙️ 其他设置")
        other_layout = QVBoxLayout(other_group)

        self.auto_save_log_cb = QCheckBox("自动保存处理日志")
        self.auto_save_log_cb.setChecked(self.config.get('auto_save_log', True))
        other_layout.addWidget(self.auto_save_log_cb)

        self.fast_mode_cb = QCheckBox("启用快速模式（向量化优化）")
        self.fast_mode_cb.setChecked(self.config.get('use_fast_mode', True))
        other_layout.addWidget(self.fast_mode_cb)

        layout.addWidget(other_group)

        # 最近文件
        recent_group = QGroupBox("📁 最近使用的文件")
        recent_layout = QVBoxLayout(recent_group)
        self.recent_list = QTextEdit()
        self.recent_list.setReadOnly(True)
        self.recent_list.setMaximumHeight(150)
        recent_files = self.config.get('recent_files', [])
        if recent_files:
            self.recent_list.setText('\n'.join([f"{i+1}. {f}" for i, f in enumerate(recent_files)]))
        else:
            self.recent_list.setText("暂无最近文件")
        recent_layout.addWidget(self.recent_list)

        clear_recent_btn = QPushButton("清空历史记录")
        clear_recent_btn.clicked.connect(self._clear_recent_files)
        recent_layout.addWidget(clear_recent_btn)

        layout.addWidget(recent_group)

        # 按钮
        layout.addStretch()
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("💾 保存设置")
        save_btn.setObjectName("successBtn")
        save_btn.clicked.connect(self._save_settings)
        btn_layout.addWidget(save_btn)

        reset_btn = QPushButton("🔄 恢复默认")
        reset_btn.clicked.connect(self._reset_settings)
        btn_layout.addWidget(reset_btn)

        layout.addLayout(btn_layout)

        return tab

    def load_settings(self):
        """加载设置到界面"""
        # 加载上季度汇率
        self.q_prev_usd.setValue(self.config.get('prev_usd_rate', 7.21))
        self.q_prev_hkd.setValue(self.config.get('prev_hkd_rate', 0.927))
        # 加载当季度汇率
        self.q_current_usd.setValue(self.config.get('current_usd_rate', 7.21))
        self.q_current_hkd.setValue(self.config.get('current_hkd_rate', 0.927))
        # 加载年份
        self.q_year.setValue(self.config.get('year', 2025))
        self.c_year.setValue(self.config.get('year', 2025))

    def _save_settings(self):
        """保存设置"""
        self.config['prev_usd_rate'] = self.settings_prev_usd.value()
        self.config['prev_hkd_rate'] = self.settings_prev_hkd.value()
        self.config['current_usd_rate'] = self.settings_current_usd.value()
        self.config['current_hkd_rate'] = self.settings_current_hkd.value()
        self.config['auto_save_log'] = self.auto_save_log_cb.isChecked()
        self.config['use_fast_mode'] = self.fast_mode_cb.isChecked()

        ConfigManager.save_config(self.config)

        # 同步更新到季度处理页面
        self.q_prev_usd.setValue(self.config['prev_usd_rate'])
        self.q_prev_hkd.setValue(self.config['prev_hkd_rate'])
        self.q_current_usd.setValue(self.config['current_usd_rate'])
        self.q_current_hkd.setValue(self.config['current_hkd_rate'])

        QMessageBox.information(self, "成功", "设置已保存！")

    def _reset_settings(self):
        """恢复默认设置"""
        self.settings_prev_usd.setValue(7.21)
        self.settings_prev_hkd.setValue(0.927)
        self.settings_current_usd.setValue(7.21)
        self.settings_current_hkd.setValue(0.927)
        self.auto_save_log_cb.setChecked(True)
        self.fast_mode_cb.setChecked(True)

    def _clear_recent_files(self):
        """清空最近文件"""
        self.config['recent_files'] = []
        ConfigManager.save_config(self.config)
        self.recent_list.setText("暂无最近文件")
        QMessageBox.information(self, "成功", "历史记录已清空！")

    def _start_quarter(self):
        if not self.q_current_file or not self.q_prev_file:
            QMessageBox.warning(self, "警告", "请选择当季度和上季度文件！")
            return
        y, q = int(self.q_year.value()), self.q_quarter.currentIndex() + 1

        # 计算上季度(输出文件是上季度的处理结果)
        if q == 1:
            prev_q, prev_y = 4, y - 1
        else:
            prev_q, prev_y = q - 1, y

        out, _ = QFileDialog.getSaveFileName(self, "保存", f"{prev_y}Q{prev_q}_处理结果.xlsx", "Excel (*.xlsx)")
        if not out: return

        self.q_output_file = out
        self.q_btn.setEnabled(False)
        self.q_prog.setValue(0)
        self.q_log.clear()

        # 保存当前汇率和年份到配置
        self.config['prev_usd_rate'] = self.q_prev_usd.value()
        self.config['prev_hkd_rate'] = self.q_prev_hkd.value()
        self.config['current_usd_rate'] = self.q_current_usd.value()
        self.config['current_hkd_rate'] = self.q_current_hkd.value()
        self.config['year'] = y
        ConfigManager.add_recent_file(self.q_current_file, self.config)
        ConfigManager.add_recent_file(self.q_prev_file, self.config)
        ConfigManager.save_config(self.config)

        fast_mode = self.config.get('use_fast_mode', True)
        self.qw = QuarterProcessWorker(
            self.q_current_file, self.q_prev_file, out, q, y,
            self.q_current_usd.value(), self.q_current_hkd.value(),
            self.q_prev_usd.value(), self.q_prev_hkd.value(),
            fast_mode
        )
        self.qw.progress.connect(lambda v, m: (self.q_prog.setValue(v), self.q_status.setText(m)))
        self.qw.finished.connect(self._q_done)
        self.qw.log.connect(self._qlog)
        self.qw.start()

    def _q_done(self, ok, res, err):
        self.q_btn.setEnabled(True)
        if ok:
            self.q_open.setEnabled(True)

            # 自动保存日志
            if self.config.get('auto_save_log', True):
                try:
                    log_file = str(Path(res).parent / f"{Path(res).stem}_日志.txt")
                    with open(log_file, 'w', encoding='utf-8') as f:
                        f.write(self.q_log.toPlainText())
                    self._qlog(f"📝 日志已保存: {Path(log_file).name}")
                except:
                    pass

            QMessageBox.information(self, "完成", f"处理完成！\n{res}")
        else:
            self._qlog(f"❌ 失败: {res}\n{err}")
            QMessageBox.critical(self, "错误", res)

    def _start_compare(self):
        if not any(self.c_q_files.get(q) for q in ['Q1','Q2','Q3','Q4']):
            QMessageBox.warning(self, "警告", "请至少选择一个季度文件！")
            return
        if not self.c_annual_file:
            QMessageBox.warning(self, "警告", "请选择年初总表！")
            return
        
        y = int(self.c_year.value())
        out, _ = QFileDialog.getSaveFileName(self, "保存", f"{y}年度结算清单.xlsx", "Excel (*.xlsx)")
        if not out: return
        
        self.c_output_file = out
        self.c_btn.setEnabled(False)
        self.c_prog.setValue(0)
        self.c_log.clear()
        
        self.cw = YearCompareWorker(self.c_q_files.get('Q1'), self.c_q_files.get('Q2'), 
                                    self.c_q_files.get('Q3'), self.c_q_files.get('Q4'),
                                    self.c_annual_file, out, y)
        self.cw.progress.connect(lambda v, m: (self.c_prog.setValue(v), self.c_status.setText(m)))
        self.cw.finished.connect(self._c_done)
        self.cw.log.connect(self._clog)
        self.cw.start()

    def _c_done(self, ok, res, err):
        self.c_btn.setEnabled(True)
        if ok:
            self.c_open.setEnabled(True)

            # 自动保存日志
            if self.config.get('auto_save_log', True):
                try:
                    log_file = str(Path(res).parent / f"{Path(res).stem}_日志.txt")
                    with open(log_file, 'w', encoding='utf-8') as f:
                        f.write(self.c_log.toPlainText())
                    self._clog(f"📝 日志已保存: {Path(log_file).name}")
                except:
                    pass

            QMessageBox.information(self, "完成", f"对比完成！\n{res}")
        else:
            self._clog(f"❌ 失败: {res}\n{err}")
            QMessageBox.critical(self, "错误", res)

    def _qlog(self, m):
        self.q_log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {m}")
    
    def _clog(self, m):
        self.c_log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {m}")

    def _open(self, t):
        f = self.q_output_file if t == 'q' else self.c_output_file
        if f: QDesktopServices.openUrl(QUrl.fromLocalFile(f))


def main():
    for pkg in ['pandas', 'openpyxl']:
        try: __import__(pkg)
        except: print(f"请安装: pip install {pkg}"); return
    
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    font = QFont('PingFang SC' if sys.platform == 'darwin' else 'Microsoft YaHei', 10)
    app.setFont(font)
    
    win = InsuranceFeeProcessor()
    win.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
